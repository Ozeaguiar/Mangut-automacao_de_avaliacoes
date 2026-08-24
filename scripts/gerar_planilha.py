"""
Gera planilha/modelo_planilha.xlsx — o arquivo que você sobe para o Google Drive
e abre como Planilhas Google.

As colunas vêm de COLUNAS em tests/gerar_workflows.py, o MESMO lugar que
alimenta o schema dos nodes do Google Sheets. Assim a planilha e os workflows
não saem de sincronia: mudou a coluna, rode os dois geradores.

Uso:  python scripts/gerar_planilha.py
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RAIZ / "tests"))
from gerar_workflows import COLUNAS  # noqa: E402

DESTINO = RAIZ / "planilha" / "modelo_planilha.xlsx"
ULTIMA_LINHA = 2000  # alcance das fórmulas do painel

FONTE = "Arial"
AZUL = "2A78D6"
CINZA_CLARO = "F5F5F3"
AMARELO = "FFF3C4"

f_titulo = Font(name=FONTE, size=14, bold=True, color="0B0B0B")
f_cabecalho = Font(name=FONTE, size=11, bold=True, color="FFFFFF")
f_normal = Font(name=FONTE, size=11)
f_nota = Font(name=FONTE, size=10, italic=True, color="52514E")
f_metrica = Font(name=FONTE, size=20, bold=True)

preenche_cabecalho = PatternFill("solid", fgColor=AZUL)
preenche_voce = PatternFill("solid", fgColor=AMARELO)
preenche_faixa = PatternFill("solid", fgColor=CINZA_CLARO)
borda_fina = Border(*[Side(style="thin", color="D9D8D4")] * 4)

# Quem preenche cada coluna: você (na mão / vindo da loja) ou o n8n.
PREENCHIDO_POR_VOCE = {
    "vendas": {"id_venda", "data_compra", "nome", "email", "telefone", "produto"},
    "respostas": set(),
    "analise": set(),
}


def montar_aba(wb, nome: str, largura_padrao: int = 18):
    ws = wb.create_sheet(nome)
    colunas = COLUNAS[nome]

    for i, col in enumerate(colunas, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = f_cabecalho
        c.fill = preenche_cabecalho
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = borda_fina
        largura = {"observacoes": 52, "resumo": 60, "temas_principais": 42,
                   "acoes_sugeridas": 42, "email": 26, "produto": 28, "token": 28,
                   "respondido_em": 22, "convite_enviado_em": 22, "analisado_em": 22}
        ws.column_dimensions[get_column_letter(i)].width = largura.get(col, largura_padrao)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    return ws


def aba_vendas(wb):
    ws = montar_aba(wb, "vendas")
    hoje = datetime.now()

    exemplos = [
        ("V-1001", (hoje - timedelta(days=3)).strftime("%Y-%m-%d"), "Gustavo Silva",
         "troque-pelo-seu@email.com", "11987654321", "Fone Bluetooth Aurora X2"),
        ("V-1002", (hoje - timedelta(days=2)).strftime("%Y-%m-%d"), "Marina Duarte",
         "troque-pelo-seu@email.com", "21996541230", "Cafeteira Prensa Nórdica"),
        ("V-1003", (hoje - timedelta(days=2)).strftime("%Y-%m-%d"), "Rafael Nunes",
         "troque-pelo-seu@email.com", "31998887766", "Luminária de Arco Bruma"),
    ]
    for r, linha in enumerate(exemplos, start=2):
        for c, valor in enumerate(linha, start=1):
            cel = ws.cell(row=r, column=c, value=valor)
            cel.font = f_normal
            cel.fill = preenche_voce
            cel.border = borda_fina
            cel.alignment = Alignment(vertical="center")

    aviso = ws.cell(row=6, column=1,
                    value="↑ Linhas de exemplo (fundo amarelo). Troque o e-mail pelo seu para testar "
                          "e apague o que não usar. As colunas token e convite_enviado_em são "
                          "preenchidas pelo n8n — deixe em branco.")
    aviso.font = f_nota
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(COLUNAS["vendas"]))
    return ws


def aba_painel(wb):
    """Painel com fórmulas ao vivo sobre a aba `respostas`."""
    ws = wb.create_sheet("painel", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col, larg in (("B", 24), ("C", 16), ("D", 24), ("E", 16), ("F", 24), ("G", 16)):
        ws.column_dimensions[col].width = larg

    t = ws["B2"]
    t.value = "Painel de satisfação"
    t.font = f_titulo
    ws["B3"].value = "Atualiza sozinho conforme as respostas chegam na aba `respostas`."
    ws["B3"].font = f_nota

    fim = ULTIMA_LINHA
    metricas = [
        ("B5", "Respostas recebidas", f"=COUNT(respostas!H2:H{fim})", "0"),
        ("D5", "Nota média", f"=IFERROR(AVERAGE(respostas!H2:H{fim}),0)", "0.0"),
        ("F5", "NPS", f'=IFERROR(ROUND((COUNTIF(respostas!H2:H{fim},">=9")'
                      f'-COUNTIF(respostas!H2:H{fim},"<=6"))/COUNT(respostas!H2:H{fim})*100,0),0)', "0"),
    ]
    for ancora, rotulo, formula, formato in metricas:
        col = ancora[0]
        linha = int(ancora[1:])
        r = ws[f"{col}{linha}"]
        r.value = rotulo
        r.font = Font(name=FONTE, size=10, bold=True, color="52514E")
        v = ws[f"{col}{linha + 1}"]
        v.value = formula
        v.font = f_metrica
        v.number_format = formato

    ws["B8"].value = "Distribuição das notas"
    ws["B8"].font = Font(name=FONTE, size=12, bold=True)

    distribuicao = [
        ("Promotores (9–10)", f'=COUNTIF(respostas!H2:H{fim},">=9")'),
        ("Neutros (7–8)", f'=COUNTIFS(respostas!H2:H{fim},">=7",respostas!H2:H{fim},"<=8")'),
        ("Detratores (0–6)", f'=COUNTIF(respostas!H2:H{fim},"<=6")'),
    ]
    for i, (rotulo, formula) in enumerate(distribuicao):
        linha = 9 + i
        ws[f"B{linha}"].value = rotulo
        ws[f"B{linha}"].font = f_normal
        ws[f"C{linha}"].value = formula
        ws[f"C{linha}"].font = f_normal
        ws[f"C{linha}"].number_format = "0"
        ws[f"D{linha}"].value = f"=IFERROR(C{linha}/$C$13,0)"
        ws[f"D{linha}"].font = f_normal
        ws[f"D{linha}"].number_format = "0.0%"
        for col in "BCD":
            ws[f"{col}{linha}"].border = borda_fina
            if i % 2 == 0:
                ws[f"{col}{linha}"].fill = preenche_faixa

    ws["B13"].value = "Total"
    ws["B13"].font = Font(name=FONTE, size=11, bold=True)
    ws["C13"].value = "=SUM(C9:C11)"
    ws["C13"].font = Font(name=FONTE, size=11, bold=True)
    ws["C13"].number_format = "0"

    ws["B15"].value = "Sentimento (preenchido pelo workflow 3)"
    ws["B15"].font = Font(name=FONTE, size=12, bold=True)
    for i, s in enumerate(["positivo", "neutro", "negativo"]):
        linha = 16 + i
        ws[f"B{linha}"].value = s.capitalize()
        ws[f"B{linha}"].font = f_normal
        ws[f"C{linha}"].value = f'=COUNTIF(respostas!K2:K{fim},"{s}")'
        ws[f"C{linha}"].font = f_normal
        ws[f"C{linha}"].number_format = "0"

    ws["B20"].value = ("NPS = % de promotores − % de detratores. Varia de −100 a +100; "
                       "acima de 50 costuma ser considerado ótimo.")
    ws["B20"].font = f_nota
    return ws


def aba_leiame(wb):
    ws = wb.create_sheet("leia-me", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 96

    ws["B2"].value = "Como usar esta planilha"
    ws["B2"].font = f_titulo

    blocos = [
        ("", ""),
        ("O que você preenche", "Só a aba `vendas`, e só as colunas com fundo amarelo: id_venda, "
                                "data_compra, nome, email, telefone, produto."),
        ("O que o n8n preenche", "`vendas.token` e `vendas.convite_enviado_em` (workflow 1); a aba "
                                 "`respostas` inteira (workflow 2); a aba `analise` e as colunas "
                                 "sentimento/temas/analisado_em (workflow 3)."),
        ("Não renomeie colunas", "Os nodes do Google Sheets casam pelo NOME do cabeçalho. Renomear "
                                 "uma coluna quebra o workflow com erro de schema."),
        ("Não apague a linha 1", "É a linha de cabeçalho de cada aba."),
        ("`painel`", "Fórmulas ao vivo sobre a aba `respostas`. Não precisa mexer."),
        ("", ""),
        ("Abas", "leia-me · painel · vendas · respostas · analise"),
        ("", ""),
        ("Para começar", "1. Suba este arquivo no Google Drive e abra como Planilhas Google.\n"
                         "2. Na aba `vendas`, troque os e-mails de exemplo pelo seu.\n"
                         "3. Copie o ID da planilha da URL e cole no node `Configuração` dos três workflows.\n"
                         "4. Rode o workflow 1 manualmente: o convite chega no seu e-mail."),
    ]
    linha = 4
    for rotulo, texto in blocos:
        if rotulo:
            ws[f"B{linha}"].value = rotulo
            ws[f"B{linha}"].font = Font(name=FONTE, size=11, bold=True)
            ws[f"B{linha}"].alignment = Alignment(vertical="top")
            ws[f"C{linha}"].value = texto
            ws[f"C{linha}"].font = f_normal
            ws[f"C{linha}"].alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[linha].height = 15 * (1 + texto.count("\n") + len(texto) // 95)
        linha += 1

    legenda = ws[f"B{linha + 1}"]
    legenda.value = "Legenda:"
    legenda.font = Font(name=FONTE, size=10, bold=True)
    ws[f"C{linha + 1}"].value = "célula amarela = você preenche   ·   célula branca = o n8n preenche"
    ws[f"C{linha + 1}"].font = f_nota
    ws[f"B{linha + 2}"].fill = preenche_voce
    ws[f"B{linha + 2}"].border = borda_fina
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    aba_vendas(wb)
    montar_aba(wb, "respostas")
    montar_aba(wb, "analise")
    aba_painel(wb)
    aba_leiame(wb)

    # ordem final das abas
    wb._sheets = [wb["leia-me"], wb["painel"], wb["vendas"], wb["respostas"], wb["analise"]]
    wb.active = 0

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DESTINO)
    print(f"Planilha gerada em {DESTINO}")
    print("Abas:", [s.title for s in wb._sheets])


if __name__ == "__main__":
    main()
